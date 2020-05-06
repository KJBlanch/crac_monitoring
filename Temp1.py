#4291
#5028

#Above values are the machine ID's required to ping temp/humidity. DO NOT DELETE.
#This will run by itself, generate new workbooks at midnight, name them after the current date, and hopefully won't
#create any memory leaks.

#Any issues, contact Kris.



#import libraries
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import requests
import time

#open and determine which workbook is required
print("All libraries okay")
dateval = datetime.date.today()
print(dateval)
dest_filename = ('Crac Temp Log (%s).xlsx' % dateval)
print (dest_filename)
try:
    workbook = load_workbook(filename = dest_filename)
except:
    workbook = Workbook()
ws = workbook.active
ws.title = "Temp CRAC"


#Determind which row to start writing in
rowval = 1
for cell in ws['A']:
    if cell.value is None:
        rowval = int(cell.row)
        print(rowval)
        break
else:
    rowval = int(cell.row + 1)
    print(rowval)

colval = 1
checked = False

#start the log!
print("Workbook Okay. Starting Log")
print(ws.title)


try:
    print("Close when no longer required. Will resume from next empty cell")
    while True:
        timestamp = datetime.datetime.now().minute
        if timestamp % 2 == 0 and checked == False:

            #this creates a new book at midnight
            if dateval != datetime.date.today():
                dateval = datetime.date.today()
                dest_filename = ('Crac Temp Log (%s).xlsx' % dateval)
                workbook = Workbook()
                ws = workbook.active
                ws.title = "Temp CRAC"
                rowval = 1

            #ping the temp/humidity and write to the cells.
            r = requests.get('http://10.10.x.xx/httpGetSet/httpGet.htm?devId=0&Value4291=vel~pnt~4291&')
            val1 = r.text
            val1 = val1[-5:-1]
            r = requests.get('http://10.10.x.xx/httpGetSet/httpGet.htm?devId=0&Value5028=vel~pnt~5028&')
            val2 = r.text
            val2 = val2[-5:-1]
            datetimeval = str(datetime.datetime.now())
            a = ws.cell(row=rowval, column=colval, value=datetimeval)
            b = ws.cell(row=rowval, column=colval+1, value=val1)
            c = ws.cell(row=rowval, column=colval+2, value=val2)
            rowval += 1
            checked = True
            print(timestamp, "logged", val1, val2)


            workbook.save(filename = dest_filename)
        if timestamp % 2 != 0:
            checked = False
        time.sleep(20)
except KeyboardInterrupt:
    print ('Stopped')
    pass

print("closing workbook and ending program")
workbook.close()
sys.exit()

